import base64
import hashlib
import hmac
import html
import io
import json
import os
import secrets as token_secrets
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openai import OpenAI
from PIL import Image

load_dotenv()

APP_TITLE = "AI仕訳アシスタント（日本の中小企業・個人事業主向け）"
DEFAULT_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o")
CUSTOMER_DB_PATH = Path(os.getenv("CUSTOMER_DB_PATH", "customer_accounts.json"))
LEDGER_STORAGE_DIR = Path(os.getenv("LEDGER_STORAGE_DIR", "customer_ledgers"))
PLAN_CONFIG = {
    "free": {"label": "無料プラン", "limit": 1},
    "starter": {"label": "スタータープラン", "limit": 5},
    "business": {"label": "ビジネスプラン", "limit": 20},
    "accounting_firm": {"label": "会計事務所プラン", "limit": 50},
}

ACCOUNT_CODE_TABLE = [
    ("110", "現金"),
    ("120", "当座預金"),
    ("130", "普通預金"),
    ("150", "受取手形"),
    ("155", "売掛金"),
    ("170", "繰越商品"),
    ("180", "未収入金"),
    ("185", "仮払消費税"),
    ("210", "建物"),
    ("220", "備品"),
    ("240", "減価償却累計額"),
    ("310", "支払手形"),
    ("315", "買掛金"),
    ("320", "預り金"),
    ("325", "未払費用"),
    ("327", "未払消費税"),
    ("330", "仮受消費税"),
    ("360", "長期借入金"),
    ("410", "資本金"),
    ("420", "前期繰越損益"),
    ("510", "売上高"),
    ("520", "売上値引"),
    ("530", "受取利息"),
    ("610", "仕入高"),
    ("620", "給料手当"),
    ("630", "法定福利費"),
    ("640", "旅費交通費"),
    ("650", "通信費"),
    ("660", "水道光熱費"),
    ("670", "広告宣伝費"),
    ("680", "接待交際費"),
    ("690", "会議費"),
    ("700", "福利厚生費"),
    ("710", "外注費"),
    ("720", "支払手数料"),
    ("730", "支払報酬料"),
    ("740", "地代家賃"),
    ("750", "租税公課"),
    ("760", "保険料"),
    ("770", "消耗品費"),
    ("780", "荷造運賃"),
    ("790", "減価償却費"),
    ("800", "雑費"),
]

CORE_ACCOUNTING_SHEETS = ["仕訳帳", "科目マスタ", "試算表", "PL", "BS"]
OPENING_BALANCE_COLUMNS = ["コード", "科目", "分類", "帳票", "開始残高", "メモ"]

PLAN_TEXT_EXAMPLES = {
    "free": "例：2026年6月10日、法人カードでAmazon Japanへ11,000円支払い。キーボードとマウスを購入。消費税10%、適格請求書取得済み。",
    "starter": "\n".join([
        "例：1行に1取引ずつ入力できます。",
        "2026年6月10日、法人カードでAmazon Japanへ11,000円支払い。キーボードとマウスを購入。消費税10%。",
        "2026年6月11日、現金でローソンへ680円支払い。会議用のお茶を購入。",
        "2026年6月12日、普通預金からNTTへ8,800円支払い。会社携帯料金。",
    ]),
    "business": "\n".join([
        "例：月次の経費をまとめて入力できます。1行1取引で記載してください。",
        "2026年6月10日、法人カードでAmazon Japanへ11,000円支払い。PC周辺機器を購入。消費税10%。",
        "2026年6月11日、JRへ1,240円支払い。営業訪問の交通費。",
        "2026年6月12日、Zoomへ2,200円支払い。オンライン会議ツール利用料。",
        "2026年6月13日、カフェで1,500円支払い。取引先との打合せ。",
    ]),
    "accounting_firm": "\n".join([
        "例：顧問先ごとの明細をまとめて入力できます。1行1取引で記載してください。",
        "A社 2026年6月10日、Amazon Japan、11,000円、法人カード、PC周辺機器、消費税10%。",
        "A社 2026年6月11日、JR、1,240円、営業交通費。",
        "B社 2026年6月12日、NTT、8,800円、普通預金、通信費。",
        "B社 2026年6月13日、飲食店、9,900円、法人カード、取引先接待。",
    ]),
}

PAGE_CSS = """
<style>
    .stApp {
        background: #0b1014;
        color: #e5edf5;
    }
    .block-container {
        max-width: 1040px;
        padding-top: 4.25rem;
        padding-bottom: 3rem;
    }
    .app-header {
        border: 1px solid #263340;
        border-radius: 8px;
        background: #101820;
        padding: 0.8rem 1rem 0.85rem 1rem;
        margin: 0.2rem auto 0.9rem auto;
        max-width: 760px;
        text-align: center;
    }
    .app-title {
        color: #f8fafc;
        font-size: 1.34rem;
        line-height: 1.55;
        font-weight: 760;
        margin: 0;
        padding: 0;
    }
    .app-subtitle {
        color: #a8b3c2;
        font-size: 0.9rem;
        line-height: 1.45;
        margin: 0.1rem 0 0 0;
    }
    .login-shell {
        max-width: 980px;
        margin: 1.25rem auto 0 auto;
    }
    h1, h2, h3, h4, h5, h6,
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li {
        color: #e5edf5;
    }
    div[data-testid="stMetric"] {
        background: #121a21;
        border: 1px solid #2a3642;
        border-radius: 8px;
        padding: 0.55rem 0.75rem;
        box-shadow: none;
        min-height: 72px;
    }
    div[data-testid="stMetric"] label {
        color: #a8b3c2 !important;
        font-weight: 650;
        font-size: 0.78rem !important;
    }
    div[data-testid="stMetricValue"] {
        color: #f8fafc !important;
        font-weight: 750;
        font-size: 1.35rem !important;
    }
    div[data-testid="stMetricValue"] > div {
        color: #f8fafc !important;
        font-size: 1.35rem !important;
    }
    section[data-testid="stSidebar"] {
        background: #10161d;
        border-right: 1px solid #24303a;
    }
    section[data-testid="stSidebar"] * {
        color: #e5edf5;
    }
    .app-panel {
        border: 1px solid #2a3642;
        border-radius: 8px;
        padding: 0.85rem 1rem;
        background: #121a21;
        margin-bottom: 1rem;
    }
    .app-panel-muted {
        border: 1px solid #263340;
        border-radius: 8px;
        padding: 0.85rem 1rem;
        background: #0f171e;
        margin-bottom: 1rem;
    }
    .eyebrow {
        color: #7dd3fc;
        font-size: 0.82rem;
        font-weight: 700;
        letter-spacing: 0;
        margin-bottom: 0.25rem;
    }
    .panel-title {
        color: #f8fafc;
        font-size: 1rem;
        font-weight: 700;
        margin-bottom: 0.2rem;
    }
    .panel-copy {
        color: #cbd5e1;
        font-size: 0.94rem;
        line-height: 1.55;
        margin: 0;
    }
    .plan-pill {
        display: inline-block;
        border: 1px solid #2dd4bf;
        border-radius: 999px;
        padding: 0.2rem 0.55rem;
        color: #ccfbf1;
        background: #102a2a;
        font-size: 0.82rem;
        font-weight: 600;
    }
    div[data-testid="stHorizontalBlock"]:has(div[data-testid="stMetric"]) {
        gap: 0.75rem;
    }
    div[data-testid="stAlert"] {
        border-radius: 8px;
    }
    div[data-testid="stDataFrame"],
    div[data-testid="stDataEditor"] {
        border: 1px solid #2a3642;
        border-radius: 8px;
        overflow: hidden;
    }
    textarea,
    input {
        color: #f8fafc !important;
    }
</style>
"""

EXCEL_COLUMNS = [
    "取引日", "証憑日付", "取引先", "摘要", "借方勘定科目", "借方補助科目", "借方金額", "借方税区分",
    "貸方勘定科目", "貸方補助科目", "貸方金額", "貸方税区分", "税込金額", "税抜金額", "消費税率",
    "消費税額", "支払方法", "インボイス登録番号", "証憑種類", "ステータス", "信頼度", "確認事項", "元ファイル名",
]

MONTHLY_CHECKLIST_ITEMS = [
    ("証憑回収", "領収書・請求書・カード明細・銀行明細が揃っているか確認"),
    ("預金照合", "普通預金・カード・電子マネー残高と帳簿残高を照合"),
    ("売掛金/買掛金", "未回収・未払の残高と入金/支払予定を確認"),
    ("未払/前払", "家賃・通信費・サブスクなど期間対応が必要な費用を確認"),
    ("給与/源泉", "給与・社会保険・源泉所得税の計上漏れを確認"),
    ("消費税区分", "課税・非課税・不課税・対象外とインボイス番号を確認"),
    ("固定資産候補", "10万円以上または資産性のある支出を確認"),
    ("確認事項", "AIが出した確認事項を原本・税理士へ確認"),
]

YEARLY_CHECKLIST_ITEMS = [
    ("年間証憑整理", "年間の証憑・契約書・明細を保管し、不足を確認"),
    ("棚卸", "商品・材料・仕掛品がある場合は期末棚卸を確認"),
    ("減価償却", "固定資産台帳、取得・売却・除却、償却費を確認"),
    ("決算整理", "未払費用・前払費用・未収収益・前受収益を確認"),
    ("売掛/買掛残高", "期末残高と入金/支払予定の整合性を確認"),
    ("貸倒/不良債権", "長期未回収の債権がある場合は回収可能性を確認"),
    ("消費税", "課税売上割合、インボイス、簡易/原則など申告方式を確認"),
    ("税務申告", "法人税・所得税・地方税・源泉所得税などを専門家へ確認"),
]

ACCOUNTING_SCHEMA = {
    "type": "object",
    "additionalProperties": False,
    "properties": {
        "transactions": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "properties": {
                    "transaction_date": {"type": "string"},
                    "evidence_date": {"type": "string"},
                    "vendor": {"type": "string"},
                    "description": {"type": "string"},
                    "gross_amount": {"type": "integer"},
                    "net_amount": {"type": "integer"},
                    "tax_amount": {"type": "integer"},
                    "tax_rate": {"type": "string"},
                    "debit_account": {"type": "string"},
                    "debit_sub_account": {"type": "string"},
                    "debit_amount": {"type": "integer"},
                    "debit_tax_category": {"type": "string"},
                    "credit_account": {"type": "string"},
                    "credit_sub_account": {"type": "string"},
                    "credit_amount": {"type": "integer"},
                    "credit_tax_category": {"type": "string"},
                    "payment_method": {"type": "string"},
                    "invoice_registration_number": {"type": "string"},
                    "evidence_type": {"type": "string"},
                    "status": {"type": "string"},
                    "confidence": {"type": "number"},
                    "review_notes": {"type": "array", "items": {"type": "string"}},
                },
                "required": [
                    "transaction_date", "evidence_date", "vendor", "description", "gross_amount", "net_amount", "tax_amount", "tax_rate",
                    "debit_account", "debit_sub_account", "debit_amount", "debit_tax_category", "credit_account", "credit_sub_account",
                    "credit_amount", "credit_tax_category", "payment_method", "invoice_registration_number", "evidence_type",
                    "status", "confidence", "review_notes",
                ],
            },
        }
    },
    "required": ["transactions"],
}

SYSTEM_PROMPT = """
あなたは日本の中小企業・個人事業主向けのAI経理・仕訳アシスタントです。

目的：
領収書・請求書・レシート・クレジットカード明細・銀行明細の画像またはテキストから、日本の小規模事業者が確認しやすい仕訳候補を作成してください。

重要ルール：
1. 出力は指定JSONスキーマに厳密に従うこと。
2. 不明な項目は必ず「要確認」とすること。推測で断定しないこと。
3. 金額は円の整数で出力すること。
4. 借方金額と貸方金額は必ず一致させること。
5. 税込金額・税抜金額・消費税額を可能な範囲で分離すること。
6. 日本の消費税区分を意識すること。
7. インボイス登録番号（T+13桁）が見える場合は必ず抽出すること。
8. 画像の読み取りが不鮮明な場合は review_notes にその旨を書くこと。
9. 最終的な税務判断は税理士確認が必要であることを review_notes に含めること。
10. 複数取引が明確に分かれる場合は transactions に複数行で出力すること。

よく使う勘定科目：
現金、普通預金、売掛金、買掛金、未払金、未収入金、前払費用、売上高、仕入高、消耗品費、
事務用品費、旅費交通費、通信費、水道光熱費、広告宣伝費、接待交際費、会議費、福利厚生費、
外注費、支払手数料、支払報酬料、地代家賃、租税公課、保険料、給料手当、法定福利費、
工具器具備品、ソフトウェア、減価償却費、雑費、仮払消費税等、仮受消費税等。

税区分候補：
課税仕入10%、課税仕入8%、課税売上10%、課税売上8%、非課税、不課税、免税、対象外、要確認。

判断方針：
- 事務用品・PC周辺機器・少額備品：消耗品費
- 10万円以上など固定資産の可能性があるもの：工具器具備品またはソフトウェアを検討し、review_notes に確認事項を書く
- 電車・タクシー・出張関連：旅費交通費
- 顧客との会食：接待交際費
- 社内会議・打合せの飲食：会議費
- 従業員向け飲食・福利厚生：福利厚生費
- 支払方法がクレジットカードの場合、貸方は通常「未払金」
- 現金払いの場合、貸方は「現金」
"""

def get_api_key() -> str:
    try:
        key = st.secrets.get("OPENAI_API_KEY")
        if key:
            return key
    except Exception:
        pass
    return os.getenv("OPENAI_API_KEY", "")


def get_config_value(name: str, default: str = "") -> str:
    try:
        value = st.secrets.get(name)
        if value:
            return str(value)
    except Exception:
        pass
    return os.getenv(name, default)


def escape_html(value: str) -> str:
    return html.escape(str(value), quote=True)


def hash_password(password: str, salt: str | None = None) -> str:
    salt = salt or token_secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), 120000)
    return f"pbkdf2_sha256${salt}${digest.hex()}"


def verify_password(password: str, profile: dict) -> bool:
    password_hash = str(profile.get("password_hash", ""))
    if password_hash.startswith("pbkdf2_sha256$"):
        try:
            _, salt, expected = password_hash.split("$", 2)
        except ValueError:
            return False
        actual = hash_password(password, salt).split("$", 2)[2]
        return hmac.compare_digest(actual, expected)

    plain_password = str(profile.get("password", ""))
    return bool(plain_password) and hmac.compare_digest(password, plain_password)


def load_registered_customer_accounts() -> dict[str, dict]:
    if not CUSTOMER_DB_PATH.exists():
        return {}

    try:
        with CUSTOMER_DB_PATH.open("r", encoding="utf-8") as file:
            data = json.load(file)
    except (OSError, json.JSONDecodeError):
        return {}

    return data if isinstance(data, dict) else {}


def save_registered_customer_accounts(customers: dict[str, dict]) -> None:
    CUSTOMER_DB_PATH.write_text(
        json.dumps(customers, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def load_customer_accounts() -> dict[str, dict]:
    customers: dict[str, dict] = load_registered_customer_accounts()

    raw_customers = get_config_value("CUSTOMER_ACCOUNTS")
    if raw_customers:
        try:
            parsed = json.loads(raw_customers)
            if isinstance(parsed, dict):
                customers.update(parsed)
        except json.JSONDecodeError:
            pass

    try:
        secret_customers = st.secrets.get("customers", {})
        customers.update({str(username): dict(profile) for username, profile in secret_customers.items()})
    except Exception:
        pass

    normalized: dict[str, dict] = {}
    for username, profile in customers.items():
        if not isinstance(profile, dict):
            continue

        plan_key = str(profile.get("plan", "free")).strip()
        password = str(profile.get("password", ""))
        password_hash = str(profile.get("password_hash", ""))
        if not username or plan_key not in PLAN_CONFIG:
            continue
        if not password and not password_hash:
            continue

        normalized[str(username).strip()] = {
            "password": password,
            "password_hash": password_hash,
            "plan": plan_key,
            "name": str(profile.get("name", username)).strip() or str(username),
        }

    return normalized


def authenticate_customer(username: str, password: str) -> dict | None:
    username = (username or "").strip()
    password = password or ""
    profile = load_customer_accounts().get(username)

    if not profile:
        return None
    if not verify_password(password, profile):
        return None

    plan_key = profile["plan"]
    return {
        "username": username,
        "name": profile["name"],
        "plan_key": plan_key,
        "plan_label": PLAN_CONFIG[plan_key]["label"],
        "transaction_limit": PLAN_CONFIG[plan_key]["limit"],
    }


def register_free_customer(username: str, password: str, name: str) -> tuple[bool, str]:
    username = (username or "").strip().lower()
    name = (name or "").strip() or username
    password = password or ""

    if not username or "@" not in username:
        return False, "メールアドレスを入力してください。"
    if len(password) < 8:
        return False, "パスワードは8文字以上で入力してください。"
    if username in load_customer_accounts():
        return False, "このメールアドレスはすでに登録されています。"

    registered_customers = load_registered_customer_accounts()
    registered_customers[username] = {
        "password_hash": hash_password(password),
        "plan": "free",
        "name": name,
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }

    try:
        save_registered_customer_accounts(registered_customers)
    except OSError:
        return False, "登録情報を保存できませんでした。管理者にお問い合わせください。"

    return True, "無料プランで登録しました。"


def get_logged_in_customer() -> dict | None:
    customer = st.session_state.get("customer")
    if isinstance(customer, dict) and customer.get("username") and customer.get("plan_key") in PLAN_CONFIG:
        return customer
    return None


def logout_customer() -> None:
    for key in ["customer", "df", "result"]:
        st.session_state.pop(key, None)
    st.rerun()


def render_login() -> None:
    upgrade_contact = get_config_value("UPGRADE_CONTACT", "上位プランをご希望の場合は管理者までお問い合わせください。")
    safe_upgrade_contact = escape_html(upgrade_contact)

    st.markdown("<div class=\"login-shell\">", unsafe_allow_html=True)
    left, right = st.columns([1, 1], gap="large")

    with left:
        st.markdown(
            """
            <div class="app-panel-muted">
                <div class="eyebrow">SMALL BUSINESS ACCOUNTING</div>
                <div class="panel-title">領収書から仕訳候補まで、確認しやすく。</div>
                <p class="panel-copy">
                    日本の中小企業・個人事業主向けに、証憑画像や取引メモから仕訳候補を作成します。
                    無料登録後は1取引ずつ試せます。
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown(
            f"""
            <div class="app-panel">
                <div class="panel-title">上位プラン</div>
                <p class="panel-copy">{safe_upgrade_contact}</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

    with right:
        login_tab, register_tab = st.tabs(["ログイン", "無料登録"])

        with login_tab:
            with st.form("customer_login_form"):
                username = st.text_input("メールアドレス / ユーザーID")
                password = st.text_input("パスワード", type="password")
                submitted = st.form_submit_button("ログイン", type="primary", width="stretch")

            if submitted:
                customer = authenticate_customer(username, password)
                if customer:
                    st.session_state["customer"] = customer
                    st.rerun()
                st.error("ユーザーIDまたはパスワードが正しくありません。")

        with register_tab:
            st.caption("登録直後は無料プランです。")
            with st.form("customer_register_form"):
                register_name = st.text_input("会社名 / お名前")
                register_email = st.text_input("メールアドレス")
                register_password = st.text_input("パスワード（8文字以上）", type="password")
                register_password_confirm = st.text_input("パスワード確認", type="password")
                registered = st.form_submit_button("無料登録して始める", type="primary", width="stretch")

            if registered:
                if register_password != register_password_confirm:
                    st.error("確認用パスワードが一致しません。")
                else:
                    success, message = register_free_customer(register_email, register_password, register_name)
                    if not success:
                        st.error(message)
                        st.stop()
                    customer = authenticate_customer(register_email, register_password)
                    if customer:
                        st.session_state["customer"] = customer
                        st.success(message)
                        st.rerun()

    st.markdown("</div>", unsafe_allow_html=True)
    st.stop()


def get_client() -> OpenAI:
    api_key = get_api_key()
    if not api_key:
        st.error("OPENAI_API_KEY が設定されていません。Streamlit Secrets または .env に設定してください。")
        st.stop()
    return OpenAI(api_key=api_key)


def image_to_data_url(uploaded_file) -> str:
    image = Image.open(uploaded_file).convert("RGB")
    image.thumbnail((1800, 1800))
    buffer = io.BytesIO()
    image.save(buffer, format="JPEG", quality=88)
    encoded = base64.b64encode(buffer.getvalue()).decode("utf-8")
    return f"data:image/jpeg;base64,{encoded}"


def parse_json_safely(text: str) -> dict:
    cleaned = (text or "").strip()
    if cleaned.startswith("```"):
        cleaned = cleaned.strip("`")
        cleaned = cleaned.replace("json\n", "", 1).replace("JSON\n", "", 1)
    return json.loads(cleaned)


def analyze_with_ai_vision(client: OpenAI, model: str, text_input: str, uploaded_file) -> dict:
    content = [
        {
            "type": "input_text",
            "text": "以下の証憑画像または取引内容から、日本会計向けの仕訳候補JSONを作成してください。\n\nユーザー入力：\n" + (text_input or "なし"),
        }
    ]
    if uploaded_file is not None:
        content.append({"type": "input_image", "image_url": image_to_data_url(uploaded_file)})

    response = client.responses.create(
        model=model,
        input=[
            {"role": "system", "content": [{"type": "input_text", "text": SYSTEM_PROMPT}]},
            {"role": "user", "content": content},
        ],
        text={
            "format": {
                "type": "json_schema",
                "name": "japanese_accounting_entries",
                "schema": ACCOUNTING_SCHEMA,
                "strict": True,
            }
        },
    )
    return parse_json_safely(response.output_text)


def transactions_to_dataframe(result: dict, file_name: str = "") -> pd.DataFrame:
    rows = []
    for item in result.get("transactions", []):
        rows.append(
            {
                "取引日": item.get("transaction_date", "要確認"),
                "証憑日付": item.get("evidence_date", "要確認"),
                "取引先": item.get("vendor", "要確認"),
                "摘要": item.get("description", "要確認"),
                "借方勘定科目": item.get("debit_account", "要確認"),
                "借方補助科目": item.get("debit_sub_account", ""),
                "借方金額": item.get("debit_amount", 0),
                "借方税区分": item.get("debit_tax_category", "要確認"),
                "貸方勘定科目": item.get("credit_account", "要確認"),
                "貸方補助科目": item.get("credit_sub_account", ""),
                "貸方金額": item.get("credit_amount", 0),
                "貸方税区分": item.get("credit_tax_category", "要確認"),
                "税込金額": item.get("gross_amount", 0),
                "税抜金額": item.get("net_amount", 0),
                "消費税率": item.get("tax_rate", "要確認"),
                "消費税額": item.get("tax_amount", 0),
                "支払方法": item.get("payment_method", "要確認"),
                "インボイス登録番号": item.get("invoice_registration_number", "要確認"),
                "証憑種類": item.get("evidence_type", "要確認"),
                "ステータス": item.get("status", "確認待ち"),
                "信頼度": item.get("confidence", 0),
                "確認事項": " / ".join(item.get("review_notes", [])),
                "元ファイル名": file_name,
            }
        )
    return pd.DataFrame(rows, columns=EXCEL_COLUMNS)


def prepare_closing_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    closing_df = df.copy()
    if closing_df.empty:
        return closing_df

    closing_df["取引日_dt"] = pd.to_datetime(closing_df["取引日"], errors="coerce")
    closing_df["月"] = closing_df["取引日_dt"].dt.strftime("%Y-%m").fillna("日付要確認")
    closing_df["年"] = closing_df["取引日_dt"].dt.strftime("%Y").fillna("日付要確認")

    for column in ["借方金額", "貸方金額", "税込金額", "税抜金額", "消費税額"]:
        closing_df[column] = pd.to_numeric(closing_df[column], errors="coerce").fillna(0).astype(int)

    closing_df["確認フラグ"] = closing_df["確認事項"].astype(str).str.len() > 0
    return closing_df


def build_monthly_closing_report(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    closing_df = prepare_closing_dataframe(df)
    if closing_df.empty:
        empty_summary = pd.DataFrame(columns=["月", "取引件数", "税込合計", "税抜合計", "消費税合計", "確認待ち件数"])
        empty_accounts = pd.DataFrame(columns=["月", "借方勘定科目", "税込合計", "消費税合計", "取引件数"])
        empty_taxes = pd.DataFrame(columns=["月", "借方税区分", "税込合計", "消費税合計", "取引件数"])
        return empty_summary, empty_accounts, empty_taxes

    monthly_summary = (
        closing_df.groupby("月", dropna=False)
        .agg(
            取引件数=("税込金額", "count"),
            税込合計=("税込金額", "sum"),
            税抜合計=("税抜金額", "sum"),
            消費税合計=("消費税額", "sum"),
            確認待ち件数=("確認フラグ", "sum"),
        )
        .reset_index()
        .sort_values("月")
    )

    account_summary = (
        closing_df.groupby(["月", "借方勘定科目"], dropna=False)
        .agg(
            税込合計=("税込金額", "sum"),
            消費税合計=("消費税額", "sum"),
            取引件数=("税込金額", "count"),
        )
        .reset_index()
        .sort_values(["月", "税込合計"], ascending=[True, False])
    )

    tax_summary = (
        closing_df.groupby(["月", "借方税区分"], dropna=False)
        .agg(
            税込合計=("税込金額", "sum"),
            消費税合計=("消費税額", "sum"),
            取引件数=("税込金額", "count"),
        )
        .reset_index()
        .sort_values(["月", "税込合計"], ascending=[True, False])
    )

    return monthly_summary, account_summary, tax_summary


def build_yearly_closing_report(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    closing_df = prepare_closing_dataframe(df)
    if closing_df.empty:
        empty_summary = pd.DataFrame(columns=["年", "取引件数", "税込合計", "税抜合計", "消費税合計", "確認待ち件数"])
        empty_accounts = pd.DataFrame(columns=["年", "借方勘定科目", "税込合計", "消費税合計", "取引件数"])
        empty_pending = pd.DataFrame(columns=EXCEL_COLUMNS)
        return empty_summary, empty_accounts, empty_pending

    yearly_summary = (
        closing_df.groupby("年", dropna=False)
        .agg(
            取引件数=("税込金額", "count"),
            税込合計=("税込金額", "sum"),
            税抜合計=("税抜金額", "sum"),
            消費税合計=("消費税額", "sum"),
            確認待ち件数=("確認フラグ", "sum"),
        )
        .reset_index()
        .sort_values("年")
    )

    account_summary = (
        closing_df.groupby(["年", "借方勘定科目"], dropna=False)
        .agg(
            税込合計=("税込金額", "sum"),
            消費税合計=("消費税額", "sum"),
            取引件数=("税込金額", "count"),
        )
        .reset_index()
        .sort_values(["年", "税込合計"], ascending=[True, False])
    )

    pending_items = closing_df[closing_df["確認フラグ"]].copy()
    pending_items = pending_items[[column for column in EXCEL_COLUMNS if column in pending_items.columns]]
    return yearly_summary, account_summary, pending_items


def build_checklist(items: list[tuple[str, str]], period_label: str) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "区分": period_label,
                "チェック項目": title,
                "確認内容": description,
                "ステータス": "未確認",
                "メモ": "",
            }
            for title, description in items
        ]
    )


def build_excel(df: pd.DataFrame, opening_balances: pd.DataFrame | None = None) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        workbook = writer.book
        add_accounting_program_sheets(workbook, df, "基本記帳・決算補助", index=0, opening_balances=opening_balances)
        apply_financial_report_format(workbook)
        for sheet in workbook.worksheets:
            set_readable_column_widths(sheet, max_width=38)
    return output.getvalue()


def estimate_text_transaction_lines(text: str) -> int:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    return len(lines)


def classify_account(account: str) -> str:
    account = str(account or "")
    asset_keywords = ["現金", "預金", "売掛", "未収", "前払", "仮払", "棚卸", "商品", "建物", "車両", "工具", "備品", "ソフトウェア"]
    liability_keywords = ["買掛", "未払", "未払金", "借入", "預り", "前受", "仮受", "未払消費税"]
    equity_keywords = ["資本金", "元入金", "利益剰余", "繰越利益", "事業主借", "事業主貸"]
    revenue_keywords = ["売上", "収益", "雑収入", "受取利息"]
    expense_keywords = [
        "仕入", "費", "損", "給料", "賃金", "旅費", "交通", "通信", "水道", "光熱", "広告", "接待",
        "会議", "福利", "外注", "手数料", "家賃", "租税", "保険", "消耗", "減価償却", "雑費",
        "支払報酬", "報酬", "荷造運賃", "運賃", "発送", "配送", "送料", "支払",
    ]

    if any(keyword in account for keyword in revenue_keywords):
        return "収益"
    if any(keyword in account for keyword in expense_keywords):
        return "費用"
    if any(keyword in account for keyword in liability_keywords):
        return "負債"
    if any(keyword in account for keyword in equity_keywords):
        return "純資産"
    if any(keyword in account for keyword in asset_keywords):
        return "資産"
    return "未分類"


def read_accounting_upload(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame(columns=EXCEL_COLUMNS)

    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        try:
            raw_df = pd.read_csv(uploaded_file, encoding="utf-8-sig")
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            raw_df = pd.read_csv(uploaded_file, encoding="cp932")
    else:
        raw_df = pd.read_excel(uploaded_file)

    return normalize_uploaded_ledger(raw_df)


def read_raw_table_upload(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame()

    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        try:
            return pd.read_csv(uploaded_file, encoding="utf-8-sig")
        except UnicodeDecodeError:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding="cp932")
    return pd.read_excel(uploaded_file)


def normalize_uploaded_ledger(raw_df: pd.DataFrame) -> pd.DataFrame:
    if raw_df.empty:
        return pd.DataFrame(columns=EXCEL_COLUMNS)

    column_aliases = {
        "取引日": ["取引日", "日付", "仕訳日", "伝票日付", "発生日"],
        "証憑日付": ["証憑日付", "証憑日", "日付", "取引日"],
        "取引先": ["取引先", "相手先", "支払先", "得意先", "摘要取引先"],
        "摘要": ["摘要", "内容", "説明", "メモ", "取引内容"],
        "借方勘定科目": ["借方勘定科目", "借方科目", "借方", "勘定科目", "科目"],
        "借方補助科目": ["借方補助科目", "借方補助", "補助科目"],
        "借方金額": ["借方金額", "借方額", "金額", "税込金額", "支出", "出金"],
        "借方税区分": ["借方税区分", "税区分", "消費税区分"],
        "貸方勘定科目": ["貸方勘定科目", "貸方科目", "貸方", "相手勘定科目"],
        "貸方補助科目": ["貸方補助科目", "貸方補助"],
        "貸方金額": ["貸方金額", "貸方額", "金額", "税込金額", "収入", "入金"],
        "貸方税区分": ["貸方税区分"],
        "税込金額": ["税込金額", "金額", "借方金額", "貸方金額"],
        "税抜金額": ["税抜金額", "本体金額"],
        "消費税率": ["消費税率", "税率"],
        "消費税額": ["消費税額", "税額"],
        "支払方法": ["支払方法", "決済方法", "口座"],
        "インボイス登録番号": ["インボイス登録番号", "登録番号"],
        "証憑種類": ["証憑種類", "証憑", "書類種別"],
        "ステータス": ["ステータス"],
        "信頼度": ["信頼度"],
        "確認事項": ["確認事項", "備考", "メモ"],
        "元ファイル名": ["元ファイル名"],
    }

    normalized = pd.DataFrame()
    for target, aliases in column_aliases.items():
        source = next((column for column in aliases if column in raw_df.columns), None)
        normalized[target] = raw_df[source] if source else ""

    for column in ["借方金額", "貸方金額", "税込金額", "税抜金額", "消費税額"]:
        normalized[column] = pd.to_numeric(normalized[column], errors="coerce").fillna(0).astype(int)

    normalized["取引日"] = normalized["取引日"].replace("", "要確認")
    missing_evidence_date = normalized["証憑日付"].astype(str).str.len() == 0
    normalized.loc[missing_evidence_date, "証憑日付"] = normalized.loc[missing_evidence_date, "取引日"]
    normalized["ステータス"] = normalized["ステータス"].replace("", "取込データ")
    normalized["確認事項"] = normalized["確認事項"].fillna("")

    missing_gross = normalized["税込金額"] == 0
    normalized.loc[missing_gross, "税込金額"] = normalized.loc[missing_gross, ["借方金額", "貸方金額"]].max(axis=1)

    missing_debit = normalized["借方金額"] == 0
    normalized.loc[missing_debit, "借方金額"] = normalized.loc[missing_debit, "税込金額"]
    missing_credit = normalized["貸方金額"] == 0
    normalized.loc[missing_credit, "貸方金額"] = normalized.loc[missing_credit, "税込金額"]

    return normalized[EXCEL_COLUMNS]


def get_customer_ledger_path(customer: dict) -> Path:
    username = str(customer.get("username", "anonymous"))
    digest = hashlib.sha256(username.encode("utf-8")).hexdigest()[:24]
    return LEDGER_STORAGE_DIR / f"{digest}.json"


def get_customer_opening_balance_path(customer: dict) -> Path:
    username = str(customer.get("username", "anonymous"))
    digest = hashlib.sha256(username.encode("utf-8")).hexdigest()[:24]
    return LEDGER_STORAGE_DIR / f"{digest}_opening_balances.json"


def load_customer_ledger(customer: dict) -> pd.DataFrame:
    path = get_customer_ledger_path(customer)
    if not path.exists():
        return pd.DataFrame(columns=EXCEL_COLUMNS)

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return pd.DataFrame(columns=EXCEL_COLUMNS)

    if not isinstance(data, list):
        return pd.DataFrame(columns=EXCEL_COLUMNS)

    df = pd.DataFrame(data)
    for column in EXCEL_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    return df[EXCEL_COLUMNS]


def save_customer_ledger(customer: dict, df: pd.DataFrame) -> None:
    LEDGER_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    clean_df = normalize_uploaded_ledger(df)
    path = get_customer_ledger_path(customer)
    path.write_text(clean_df.to_json(orient="records", force_ascii=False, indent=2), encoding="utf-8")


def append_customer_ledger(customer: dict, df: pd.DataFrame) -> int:
    if df.empty:
        return len(load_customer_ledger(customer))

    existing_df = load_customer_ledger(customer)
    merged_df = pd.concat([existing_df, normalize_uploaded_ledger(df)], ignore_index=True)
    save_customer_ledger(customer, merged_df)
    return len(merged_df)


def clear_customer_ledger(customer: dict) -> None:
    path = get_customer_ledger_path(customer)
    if path.exists():
        path.unlink()


def build_opening_balance_template(existing_df: pd.DataFrame | None = None) -> pd.DataFrame:
    existing_map = {}
    if existing_df is not None and not existing_df.empty:
        for _, row in existing_df.iterrows():
            account = str(row.get("科目", "")).strip()
            if account:
                existing_map[account] = {
                    "開始残高": safe_excel_int(row.get("開始残高", 0)),
                    "メモ": str(row.get("メモ", "") or ""),
                }

    rows = []
    seen_accounts = set()
    for row in get_account_master_rows():
        existing = existing_map.get(row["科目"], {})
        rows.append({
            "コード": row["コード"],
            "科目": row["科目"],
            "分類": row["分類"],
            "帳票": row["帳票"],
            "開始残高": existing.get("開始残高", 0),
            "メモ": existing.get("メモ", ""),
        })
        seen_accounts.add(row["科目"])

    if existing_df is not None and not existing_df.empty:
        for _, row in existing_df.iterrows():
            account = str(row.get("科目", "")).strip()
            if not account or account in seen_accounts:
                continue
            category = str(row.get("分類", "") or "").strip() or classify_account(account)
            rows.append({
                "コード": str(row.get("コード", "") or ""),
                "科目": account,
                "分類": category,
                "帳票": str(row.get("帳票", "") or "").strip() or ("PL" if category in ["収益", "費用"] else "BS"),
                "開始残高": safe_excel_int(row.get("開始残高", 0)),
                "メモ": str(row.get("メモ", "") or ""),
            })
            seen_accounts.add(account)
    return pd.DataFrame(rows, columns=OPENING_BALANCE_COLUMNS)


def normalize_opening_balance_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return build_opening_balance_template()

    normalized = df.copy()
    for column in OPENING_BALANCE_COLUMNS:
        if column not in normalized.columns:
            normalized[column] = 0 if column == "開始残高" else ""

    normalized["科目"] = normalized["科目"].astype(str).str.strip()
    normalized = normalized[normalized["科目"].astype(str).str.len() > 0].copy()
    normalized["開始残高"] = pd.to_numeric(normalized["開始残高"], errors="coerce").fillna(0).astype(int)
    for idx, row in normalized.iterrows():
        account = str(row["科目"])
        category = str(row.get("分類", "") or "").strip() or classify_account(account)
        normalized.at[idx, "分類"] = category
        normalized.at[idx, "帳票"] = str(row.get("帳票", "") or "").strip() or ("PL" if category in ["収益", "費用"] else "BS")

    return normalized[OPENING_BALANCE_COLUMNS]


def load_customer_opening_balances(customer: dict) -> pd.DataFrame:
    path = get_customer_opening_balance_path(customer)
    if not path.exists():
        return build_opening_balance_template()

    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return build_opening_balance_template()

    if not isinstance(data, list):
        return build_opening_balance_template()
    return build_opening_balance_template(normalize_opening_balance_df(pd.DataFrame(data)))


def save_customer_opening_balances(customer: dict, df: pd.DataFrame) -> None:
    LEDGER_STORAGE_DIR.mkdir(parents=True, exist_ok=True)
    clean_df = normalize_opening_balance_df(df)
    path = get_customer_opening_balance_path(customer)
    path.write_text(clean_df.to_json(orient="records", force_ascii=False, indent=2), encoding="utf-8")


def build_trial_balance(df: pd.DataFrame, opening_balances: pd.DataFrame | None = None) -> pd.DataFrame:
    columns = ["区分", "勘定科目", "借方合計", "貸方合計", "残高"]
    debit = pd.Series(dtype="int64", name="借方合計")
    credit = pd.Series(dtype="int64", name="貸方合計")

    if df is not None and not df.empty:
        debit = df.groupby("借方勘定科目", dropna=False)["借方金額"].sum().rename("借方合計")
        credit = df.groupby("貸方勘定科目", dropna=False)["貸方金額"].sum().rename("貸方合計")

    trial = pd.concat([debit, credit], axis=1).fillna(0).reset_index().rename(columns={"index": "勘定科目"})
    opening_df = normalize_opening_balance_df(opening_balances) if opening_balances is not None else pd.DataFrame(columns=OPENING_BALANCE_COLUMNS)
    opening_accounts = opening_df[opening_df["開始残高"] != 0][["科目", "分類", "開始残高"]].rename(columns={"科目": "勘定科目"})
    if not opening_accounts.empty:
        trial = pd.concat([trial, opening_accounts[["勘定科目"]]], ignore_index=True)

    if trial.empty:
        return pd.DataFrame(columns=columns)

    trial = trial.drop_duplicates(subset=["勘定科目"], keep="first")
    trial = trial[trial["勘定科目"].astype(str).str.len() > 0]
    trial[["借方合計", "貸方合計"]] = trial[["借方合計", "貸方合計"]].fillna(0)
    trial["借方合計"] = trial["借方合計"].astype(int)
    trial["貸方合計"] = trial["貸方合計"].astype(int)
    trial["残高"] = trial["借方合計"] - trial["貸方合計"]
    trial["区分"] = trial["勘定科目"].apply(classify_account)

    if not opening_accounts.empty:
        opening_map = dict(zip(opening_accounts["勘定科目"], opening_accounts["開始残高"]))
        for idx, row in trial.iterrows():
            opening_value = safe_excel_int(opening_map.get(row["勘定科目"], 0))
            if opening_value == 0:
                continue
            if row["区分"] in ["収益", "負債", "純資産"]:
                trial.at[idx, "残高"] = int(row["残高"]) - opening_value
            else:
                trial.at[idx, "残高"] = int(row["残高"]) + opening_value

    return trial[columns].sort_values(["区分", "勘定科目"])


def build_profit_and_loss(trial_balance: pd.DataFrame) -> pd.DataFrame:
    rows = []
    revenue = trial_balance[trial_balance["区分"] == "収益"].copy()
    expense = trial_balance[trial_balance["区分"] == "費用"].copy()

    for _, item in revenue.iterrows():
        rows.append({"区分": "収益", "科目": item["勘定科目"], "金額": int(-item["残高"])})
    for _, item in expense.iterrows():
        rows.append({"区分": "費用", "科目": item["勘定科目"], "金額": int(item["残高"])})

    statement = pd.DataFrame(rows, columns=["区分", "科目", "金額"])
    revenue_total = int(statement[statement["区分"] == "収益"]["金額"].sum()) if not statement.empty else 0
    expense_total = int(statement[statement["区分"] == "費用"]["金額"].sum()) if not statement.empty else 0
    net_income = revenue_total - expense_total
    totals = pd.DataFrame(
        [
            {"区分": "合計", "科目": "収益合計", "金額": revenue_total},
            {"区分": "合計", "科目": "費用合計", "金額": expense_total},
            {"区分": "合計", "科目": "当期利益", "金額": net_income},
        ]
    )
    return pd.concat([statement, totals], ignore_index=True)


def build_balance_sheet(trial_balance: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for category in ["資産", "負債", "純資産", "未分類"]:
        category_df = trial_balance[trial_balance["区分"] == category]
        for _, item in category_df.iterrows():
            amount = int(item["残高"])
            if category in ["負債", "純資産"]:
                amount = -amount
            rows.append({"区分": category, "科目": item["勘定科目"], "金額": amount})

    statement = pd.DataFrame(rows, columns=["区分", "科目", "金額"])
    totals = []
    for category in ["資産", "負債", "純資産", "未分類"]:
        total = int(statement[statement["区分"] == category]["金額"].sum()) if not statement.empty else 0
        totals.append({"区分": "合計", "科目": f"{category}合計", "金額": total})
    return pd.concat([statement, pd.DataFrame(totals)], ignore_index=True)


def get_account_master_rows(df: pd.DataFrame | None = None) -> list[dict]:
    rows = []
    seen_accounts = set()

    for code, account in ACCOUNT_CODE_TABLE:
        category = classify_account(account)
        rows.append({
            "コード": code,
            "科目": account,
            "分類": category,
            "帳票": "PL" if category in ["収益", "費用"] else "BS",
        })
        seen_accounts.add(account)

    if df is not None and not df.empty:
        account_columns = ["借方勘定科目", "貸方勘定科目"]
        accounts = set()
        for column in account_columns:
            if column in df.columns:
                accounts.update(str(value).strip() for value in df[column].dropna().tolist())
        for account in sorted(account for account in accounts if account and account not in seen_accounts):
            category = classify_account(account)
            rows.append({
                "コード": "",
                "科目": account,
                "分類": category,
                "帳票": "PL" if category in ["収益", "費用"] else ("要確認" if category == "未分類" else "BS"),
            })

    return rows


def create_or_replace_sheet(workbook, title: str, index: int):
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])
    return workbook.create_sheet(title, index)


def safe_excel_int(value) -> int:
    number = pd.to_numeric(value, errors="coerce")
    return 0 if pd.isna(number) else int(number)


def excel_display_width(value) -> int:
    text = str(value or "")
    if text.startswith("="):
        return 12
    width = 0
    for char in text:
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W", "A"} else 1
    return width


def set_readable_column_widths(sheet, max_width: int = 42) -> None:
    for column_idx in range(1, sheet.max_column + 1):
        column_letter = get_column_letter(column_idx)
        max_length = max(excel_display_width(cell.value) for cell in sheet[column_letter])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), max_width)
    if sheet.title in ["PL", "BS"]:
        sheet.column_dimensions["A"].width = max(sheet.column_dimensions["A"].width or 0, 24)
        sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 18)
    elif sheet.title == "仕訳帳":
        sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 34)
    elif sheet.title in ["科目マスタ", "試算表"]:
        sheet.column_dimensions["B"].width = max(sheet.column_dimensions["B"].width or 0, 24)
        if sheet.title == "科目マスタ":
            sheet.column_dimensions["F"].width = max(sheet.column_dimensions["F"].width or 0, 30)


def style_program_sheet(sheet, header_row: int = 1, freeze: str | None = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.sheet_view.showGridLines = False
    if freeze:
        sheet.freeze_panes = freeze

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = '#,##0;[Red]-#,##0;-'

    for cell in sheet[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    set_readable_column_widths(sheet, max_width=34)


def polish_statement_sheet(sheet, title: str, period_label: str) -> None:
    sheet.merge_cells("A1:B1")
    sheet["A1"] = title
    sheet["A1"].font = Font(size=15, bold=True, color="0F172A")
    sheet["A1"].fill = PatternFill("solid", fgColor="E0F2FE")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet["A2"] = period_label
    sheet["A2"].font = Font(bold=True, color="2563EB")
    sheet["A2"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 24
    sheet.row_dimensions[2].height = 20
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 18

    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=2):
        label = str(row[0].value or "")
        if label in ["収益", "費用", "資産", "負債", "純資産"]:
            for cell in row:
                cell.font = Font(bold=True, color="1F4E78")
                cell.fill = PatternFill("solid", fgColor="DBEAFE")
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def add_accounting_program_sheets(
    workbook,
    ledger_df: pd.DataFrame,
    period_label: str,
    index: int = 0,
    opening_balances: pd.DataFrame | None = None,
) -> None:
    for sheet_name in CORE_ACCOUNTING_SHEETS:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])

    ledger_sheet = workbook.create_sheet("仕訳帳", index)
    master_sheet = workbook.create_sheet("科目マスタ", index + 1)
    trial_sheet = workbook.create_sheet("試算表", index + 2)
    pl_sheet = workbook.create_sheet("PL", index + 3)
    bs_sheet = workbook.create_sheet("BS", index + 4)

    ledger_headers = ["日付", "摘要", "借方科目", "借方金額", "貸方科目", "貸方金額"]
    ledger_sheet.append(ledger_headers)
    for _, row in ledger_df.iterrows():
        ledger_sheet.append([
            row.get("取引日", ""),
            row.get("摘要", ""),
            row.get("借方勘定科目", ""),
            safe_excel_int(row.get("借方金額", 0)),
            row.get("貸方勘定科目", ""),
            safe_excel_int(row.get("貸方金額", 0)),
        ])
    for _ in range(max(100 - len(ledger_df), 20)):
        ledger_sheet.append(["", "", "", "", "", ""])

    opening_df = normalize_opening_balance_df(opening_balances) if opening_balances is not None else build_opening_balance_template()
    opening_map = {
        str(row["科目"]): {
            "開始残高": safe_excel_int(row.get("開始残高", 0)),
            "メモ": str(row.get("メモ", "") or ""),
        }
        for _, row in opening_df.iterrows()
    }

    account_rows = get_account_master_rows(ledger_df)
    seen_accounts = {row["科目"] for row in account_rows}
    for _, opening_row in opening_df.iterrows():
        account = str(opening_row.get("科目", "")).strip()
        if not account or account in seen_accounts:
            continue
        category = str(opening_row.get("分類", "") or "").strip() or classify_account(account)
        account_rows.append({
            "コード": str(opening_row.get("コード", "") or ""),
            "科目": account,
            "分類": category,
            "帳票": str(opening_row.get("帳票", "") or "").strip() or ("PL" if category in ["収益", "費用"] else "BS"),
        })
        seen_accounts.add(account)

    master_headers = ["コード", "科目", "分類", "帳票", "開始残高", "メモ"]
    master_sheet.append(master_headers)
    for row in account_rows:
        opening = opening_map.get(row["科目"], {})
        master_sheet.append([
            row["コード"],
            row["科目"],
            row["分類"],
            row["帳票"],
            opening.get("開始残高", 0),
            opening.get("メモ", ""),
        ])

    account_list_range = f"'科目マスタ'!$B$2:$B${len(account_rows) + 1}"
    validation = DataValidation(type="list", formula1=account_list_range, allow_blank=True)
    ledger_sheet.add_data_validation(validation)
    validation.add(f"C2:C{ledger_sheet.max_row}")
    validation.add(f"E2:E{ledger_sheet.max_row}")

    trial_headers = ["コード", "科目", "分類", "開始残高", "借方合計", "貸方合計", "現在残高"]
    trial_sheet.append(trial_headers)
    for row_idx, account in enumerate(account_rows, start=2):
        trial_sheet[f"A{row_idx}"] = account["コード"]
        trial_sheet[f"B{row_idx}"] = account["科目"]
        trial_sheet[f"C{row_idx}"] = f'=IFERROR(VLOOKUP(B{row_idx},科目マスタ!$B:$D,2,FALSE),"未分類")'
        trial_sheet[f"D{row_idx}"] = f'=IFERROR(INDEX(科目マスタ!$E:$E,MATCH(B{row_idx},科目マスタ!$B:$B,0)),0)'
        trial_sheet[f"E{row_idx}"] = f'=SUMIF(仕訳帳!$C:$C,B{row_idx},仕訳帳!$D:$D)'
        trial_sheet[f"F{row_idx}"] = f'=SUMIF(仕訳帳!$E:$E,B{row_idx},仕訳帳!$F:$F)'
        trial_sheet[f"G{row_idx}"] = f'=D{row_idx}+IF(OR(C{row_idx}="収益",C{row_idx}="負債",C{row_idx}="純資産"),F{row_idx}-E{row_idx},E{row_idx}-F{row_idx})'
    trial_total_row = len(account_rows) + 2
    trial_sheet[f"B{trial_total_row}"] = "合計"
    trial_sheet[f"D{trial_total_row}"] = f"=SUM(D2:D{trial_total_row - 1})"
    trial_sheet[f"E{trial_total_row}"] = f"=SUM(E2:E{trial_total_row - 1})"
    trial_sheet[f"F{trial_total_row}"] = f"=SUM(F2:F{trial_total_row - 1})"
    trial_sheet[f"G{trial_total_row}"] = f"=SUMIF(C2:C{trial_total_row - 1},\"資産\",G2:G{trial_total_row - 1})+SUMIF(C2:C{trial_total_row - 1},\"費用\",G2:G{trial_total_row - 1})-SUMIF(C2:C{trial_total_row - 1},\"負債\",G2:G{trial_total_row - 1})-SUMIF(C2:C{trial_total_row - 1},\"純資産\",G2:G{trial_total_row - 1})-SUMIF(C2:C{trial_total_row - 1},\"収益\",G2:G{trial_total_row - 1})"

    pl_sheet["A1"] = "PL 損益計算書"
    pl_sheet["B1"] = period_label
    pl_sheet["A3"] = "項目"
    pl_sheet["B3"] = "金額"
    current_row = 4
    revenue_accounts = [row["科目"] for row in account_rows if row["分類"] == "収益"]
    expense_accounts = [row["科目"] for row in account_rows if row["分類"] == "費用"]
    for account in revenue_accounts:
        pl_sheet[f"A{current_row}"] = account
        pl_sheet[f"B{current_row}"] = f'=IFERROR(INDEX(試算表!$G:$G,MATCH(A{current_row},試算表!$B:$B,0)),0)'
        current_row += 1
    revenue_total_row = current_row
    pl_sheet[f"A{revenue_total_row}"] = "収益合計"
    pl_sheet[f"B{revenue_total_row}"] = f"=SUM(B4:B{revenue_total_row - 1})"
    current_row += 2
    expense_start_row = current_row
    for account in expense_accounts:
        pl_sheet[f"A{current_row}"] = account
        pl_sheet[f"B{current_row}"] = f'=IFERROR(INDEX(試算表!$G:$G,MATCH(A{current_row},試算表!$B:$B,0)),0)'
        current_row += 1
    expense_total_row = current_row
    pl_sheet[f"A{expense_total_row}"] = "費用合計"
    pl_sheet[f"B{expense_total_row}"] = f"=SUM(B{expense_start_row}:B{expense_total_row - 1})"
    net_income_row = current_row + 1
    pl_sheet[f"A{net_income_row}"] = "当期純利益"
    pl_sheet[f"B{net_income_row}"] = f"=B{revenue_total_row}-B{expense_total_row}"

    bs_sheet["A1"] = "BS 貸借対照表"
    bs_sheet["B1"] = period_label
    bs_sheet["A3"] = "項目"
    bs_sheet["B3"] = "金額"
    current_row = 4
    bs_total_rows = {}
    for category in ["資産", "負債", "純資産"]:
        bs_sheet[f"A{current_row}"] = category
        bs_sheet[f"A{current_row}"].font = Font(bold=True, color="1F4E78")
        current_row += 1
        category_start_row = current_row
        for account in [row["科目"] for row in account_rows if row["分類"] == category]:
            bs_sheet[f"A{current_row}"] = account
            bs_sheet[f"B{current_row}"] = f'=IFERROR(INDEX(試算表!$G:$G,MATCH(A{current_row},試算表!$B:$B,0)),0)'
            current_row += 1
        if category == "純資産":
            bs_sheet[f"A{current_row}"] = "当期純利益"
            bs_sheet[f"B{current_row}"] = f"=PL!B{net_income_row}"
            current_row += 1
        total_row = current_row
        bs_sheet[f"A{total_row}"] = f"{category}合計"
        bs_sheet[f"B{total_row}"] = f"=SUM(B{category_start_row}:B{total_row - 1})"
        bs_total_rows[category] = total_row
        current_row += 2
    bs_sheet[f"A{current_row}"] = "B/S差額チェック"
    bs_sheet[f"B{current_row}"] = f"=B{bs_total_rows['資産']}-B{bs_total_rows['負債']}-B{bs_total_rows['純資産']}"

    style_program_sheet(ledger_sheet, freeze="A2")
    style_program_sheet(master_sheet, freeze="A2")
    style_program_sheet(trial_sheet, freeze="A2")
    style_program_sheet(pl_sheet, header_row=3, freeze="A4")
    style_program_sheet(bs_sheet, header_row=3, freeze="A4")
    polish_statement_sheet(pl_sheet, "PL 損益計算書", period_label)
    polish_statement_sheet(bs_sheet, "BS 貸借対照表", period_label)

    for sheet, rows in [(trial_sheet, [trial_total_row]), (pl_sheet, [revenue_total_row, expense_total_row, net_income_row])]:
        for row_idx in rows:
            for cell in sheet[row_idx]:
                cell.font = Font(bold=True, color="0F172A")
                cell.fill = PatternFill("solid", fgColor="E2E8F0")
    for row_idx in bs_total_rows.values():
        for cell in bs_sheet[row_idx]:
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in bs_sheet[current_row]:
        cell.font = Font(bold=True, color="9F1239")
        cell.fill = PatternFill("solid", fgColor="FFE4E6")


def style_excel_sheet(sheet, freeze: str = "A2") -> None:
    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True)
    border_color = "CBD5E1"
    thin_border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    sheet.sheet_view.showGridLines = False
    if freeze:
        sheet.freeze_panes = freeze

    max_row = sheet.max_row or 1
    max_column = sheet.max_column or 1

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for row in sheet.iter_rows(min_row=2, max_row=max_row, max_col=max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0;[Red]-#,##0;-'

    for column_idx in range(1, max_column + 1):
        column_letter = get_column_letter(column_idx)
        max_length = 0
        for cell in sheet[column_letter]:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def add_cover_sheet(workbook, period_label: str, index: int = 0) -> None:
    sheet = create_or_replace_sheet(workbook, "表紙", index)
    sheet.sheet_view.showGridLines = False
    sheet["B2"] = "財務報告書"
    sheet["B3"] = period_label
    sheet["B5"] = "作成日時"
    sheet["C5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet["B7"] = "含まれる帳票"
    reports = ["仕訳帳", "科目マスタ", "試算表", "PL", "BS", "精算表", "財務サマリー"]
    for idx, report in enumerate(reports, start=8):
        sheet[f"B{idx}"] = report
    sheet["B16"] = "注意"
    sheet["C16"] = "AIとルールによる参考資料です。決算・申告の最終判断は税理士へ確認してください。"

    sheet["B2"].font = Font(size=22, bold=True, color="0F172A")
    sheet["B3"].font = Font(size=14, bold=True, color="2563EB")
    for cell_ref in ["B5", "B7", "B16"]:
        sheet[cell_ref].font = Font(bold=True, color="334155")
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 72


def add_financial_summary_sheet(workbook, period_label: str, index: int = 1) -> None:
    sheet = create_or_replace_sheet(workbook, "財務サマリー", index)
    sheet.sheet_view.showGridLines = False
    sheet["B2"] = "財務サマリー"
    sheet["C2"] = period_label

    rows = [
        ("売上高", '=SUMIF(PL!A:A,"収益合計",PL!B:B)'),
        ("費用合計", '=SUMIF(PL!A:A,"費用合計",PL!B:B)'),
        ("当期利益", '=SUMIF(PL!A:A,"当期純利益",PL!B:B)'),
        ("資産合計", '=SUMIF(BS!A:A,"資産合計",BS!B:B)'),
        ("負債合計", '=SUMIF(BS!A:A,"負債合計",BS!B:B)'),
        ("純資産合計", '=SUMIF(BS!A:A,"純資産合計",BS!B:B)'),
        ("B/S差額チェック", '=C6-C7-C8'),
    ]
    sheet["B4"] = "指標"
    sheet["C4"] = "金額"
    for row_idx, (label, formula) in enumerate(rows, start=5):
        sheet[f"B{row_idx}"] = label
        sheet[f"C{row_idx}"] = formula

    sheet["E4"] = "確認ポイント"
    notes = [
        "B/S差額チェックが0に近いか確認",
        "未分類科目がある場合は科目分類を確認",
        "確認事項シートと原本証憑を照合",
        "税区分・インボイス登録番号は税理士確認",
    ]
    for row_idx, note in enumerate(notes, start=5):
        sheet[f"E{row_idx}"] = note

    sheet["B2"].font = Font(size=18, bold=True, color="0F172A")
    sheet["C2"].font = Font(size=12, bold=True, color="2563EB")
    for cell_ref in ["B4", "C4", "E4"]:
        sheet[cell_ref].fill = PatternFill("solid", fgColor="0F172A")
        sheet[cell_ref].font = Font(color="FFFFFF", bold=True)
    for row in range(5, 12):
        sheet[f"C{row}"].number_format = '#,##0;[Red]-#,##0;-'
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 18
    sheet.column_dimensions["E"].width = 48


def add_worksheet_sheet(workbook, trial_balance: pd.DataFrame, period_label: str, index: int = 2) -> None:
    sheet = create_or_replace_sheet(workbook, "精算表", index)
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"

    sheet.merge_cells("A1:I2")
    sheet["A1"] = "精　算　表"
    sheet["A1"].font = Font(size=18, bold=True, color="1E3A8A")
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["A1"].fill = PatternFill("solid", fgColor="F2DCDB")

    sheet["C3"] = "試算表"
    sheet["C4"] = f"{period_label}残高"
    sheet.merge_cells("D3:E3")
    sheet["D3"] = "修正記入"
    sheet["D4"] = "借方"
    sheet["E4"] = "貸方"
    sheet.merge_cells("F3:G3")
    sheet["F3"] = "損益計算書"
    sheet["F4"] = "借方"
    sheet["G4"] = "貸方"
    sheet.merge_cells("H3:I3")
    sheet["H3"] = "貸借対照表"
    sheet["H4"] = "借方"
    sheet["I4"] = "貸方"
    sheet["A4"] = "コード"
    sheet["B4"] = "勘定科目"

    header_fills = {
        "C3:C3": "D9EAD3",
        "D3:E3": "CFE2F3",
        "F3:G3": "D9D2E9",
        "H3:I3": "FCE5CD",
    }
    for cell_range, color in header_fills.items():
        fill = PatternFill("solid", fgColor=color)
        for row in sheet[cell_range]:
            for cell in row:
                cell.fill = fill

    trial_map = {
        str(row["勘定科目"]): {
            "区分": row["区分"],
            "残高": int(row["残高"]),
        }
        for _, row in trial_balance.iterrows()
    }

    start_row = 5
    for idx, (code, account) in enumerate(ACCOUNT_CODE_TABLE, start=start_row):
        info = trial_map.get(account, {"区分": classify_account(account), "残高": 0})
        category = info["区分"]
        balance = int(info["残高"])
        debit_balance = balance if balance > 0 else 0
        credit_balance = -balance if balance < 0 else 0

        sheet[f"A{idx}"] = code
        sheet[f"B{idx}"] = account
        sheet[f"C{idx}"] = debit_balance if debit_balance else credit_balance

        if category == "費用":
            sheet[f"F{idx}"] = debit_balance
            sheet[f"G{idx}"] = credit_balance
        elif category == "収益":
            sheet[f"F{idx}"] = debit_balance
            sheet[f"G{idx}"] = credit_balance
        elif category == "資産":
            sheet[f"H{idx}"] = debit_balance
            sheet[f"I{idx}"] = credit_balance
        elif category in ["負債", "純資産"]:
            sheet[f"H{idx}"] = debit_balance
            sheet[f"I{idx}"] = credit_balance

    total_row = start_row + len(ACCOUNT_CODE_TABLE)
    sheet[f"B{total_row}"] = "合計"
    for col in ["C", "D", "E", "F", "G", "H", "I"]:
        sheet[f"{col}{total_row}"] = f"=SUM({col}{start_row}:{col}{total_row - 1})"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in sheet.iter_rows(min_row=3, max_row=total_row, min_col=1, max_col=9):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if cell.row in [3, 4] or cell.row == total_row:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = '#,##0;[Red]-#,##0;-'

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 22
    for col in ["C", "D", "E", "F", "G", "H", "I"]:
        sheet.column_dimensions[col].width = 14


def apply_financial_report_format(workbook) -> None:
    for sheet in workbook.worksheets:
        if sheet.title not in ["表紙", "財務サマリー", "精算表", *CORE_ACCOUNTING_SHEETS]:
            style_excel_sheet(sheet)

    for sheet_name in ["損益計算書", "貸借対照表", "試算表"]:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2):
            first_value = str(row[0].value or "")
            second_value = str(row[1].value or "") if len(row) > 1 else ""
            if first_value == "合計" or second_value.endswith("合計") or second_value == "当期利益":
                for cell in row:
                    cell.font = Font(bold=True, color="0F172A")
                    cell.fill = PatternFill("solid", fgColor="E2E8F0")


def build_financial_statement_excel(
    ledger_df: pd.DataFrame,
    trial_balance: pd.DataFrame,
    profit_and_loss: pd.DataFrame,
    balance_sheet: pd.DataFrame,
    period_label: str,
    opening_balances: pd.DataFrame | None = None,
) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        workbook = writer.book
        add_accounting_program_sheets(workbook, ledger_df, period_label, index=0, opening_balances=opening_balances)
        if period_label == "年次決算":
            add_worksheet_sheet(workbook, trial_balance, period_label, index=3)
        apply_financial_report_format(workbook)

        for sheet in workbook.worksheets:
            set_readable_column_widths(sheet, max_width=38)
    return output.getvalue()


def render_bookkeeping_workspace(customer: dict, model: str, transaction_limit: int) -> None:
    overview_cols = st.columns(4)
    overview_cols[0].metric("現在のプラン", customer["plan_label"])
    overview_cols[1].metric("処理上限", f"{transaction_limit} 取引/回")
    overview_cols[2].metric("今回の結果", f"{len(st.session_state.get('df', []))} 件")
    overview_cols[3].metric("出力形式", "Excel")

    left, right = st.columns([0.82, 1.18], gap="large")

    with left:
        st.markdown(
            """
            <div class="app-panel">
                <div class="eyebrow">BASIC BOOKKEEPING</div>
                <div class="panel-title">証憑または取引内容から記帳</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        input_mode = st.radio("入力方式", ["画像アップロード", "テキスト入力"], horizontal=True)
        uploaded_files = []
        text_input = ""

        if input_mode == "画像アップロード":
            uploaded_files = st.file_uploader(
                "領収書・請求書・明細画像",
                type=["png", "jpg", "jpeg", "webp"],
                accept_multiple_files=True,
                key="bookkeeping_images",
            )
            if uploaded_files:
                if len(uploaded_files) > transaction_limit:
                    st.error(f"現在のプランでは一度に処理できる証憑画像は{transaction_limit}件までです。")
                    st.stop()
                st.success(f"{len(uploaded_files)}件の証憑をアップロードしました")
                preview_cols = st.columns(2)
                for idx, file in enumerate(uploaded_files[:4]):
                    preview_cols[idx % 2].image(file, caption=file.name, use_container_width=True)
                if len(uploaded_files) > 4:
                    st.info(f"ほか {len(uploaded_files) - 4} 件のプレビューを省略しています。")
        else:
            text_input = st.text_area(
                "取引内容",
                height=210,
                placeholder=PLAN_TEXT_EXAMPLES.get(customer["plan_key"], PLAN_TEXT_EXAMPLES["free"]),
                key="bookkeeping_text",
            )

        run = st.button("仕訳候補を作成", type="primary", width="stretch")

    if run:
        if not uploaded_files and not text_input.strip():
            st.warning("画像をアップロードするか、取引内容を入力してください。")
            return
        if not uploaded_files:
            estimated_lines = estimate_text_transaction_lines(text_input)
            if estimated_lines > transaction_limit:
                st.error(
                    f"現在のプランでは一度に処理できる取引は{transaction_limit}件までです。"
                    f"入力は{estimated_lines}行あります。取引数を減らすか、上位プランをご利用ください。"
                )
                return
        client = get_client()
        with st.spinner("AI Visionで証憑を解析しています..."):
            try:
                all_results = []
                all_dfs = []
                if uploaded_files:
                    progress = st.progress(0)
                    for idx, file in enumerate(uploaded_files, start=1):
                        result = analyze_with_ai_vision(client, model, "", file)
                        df = transactions_to_dataframe(result, file.name)
                        all_results.append({"file_name": file.name, "result": result})
                        all_dfs.append(df)
                        progress.progress(idx / len(uploaded_files))
                else:
                    result = analyze_with_ai_vision(client, model, text_input, None)
                    transaction_count = len(result.get("transactions", []))
                    if transaction_count > transaction_limit:
                        st.error(
                            f"現在のプランでは一度に処理できる取引は{transaction_limit}件までです。"
                            f"入力内容は{transaction_count}件として解析されました。取引数を減らすか、上位プランをご利用ください。"
                        )
                        return
                    df = transactions_to_dataframe(result, "テキスト入力")
                    all_results.append({"file_name": "テキスト入力", "result": result})
                    all_dfs.append(df)
                merged_df = pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame(columns=EXCEL_COLUMNS)
                if len(merged_df) > transaction_limit:
                    st.error(
                        f"現在のプランでは一度に処理できる取引は{transaction_limit}件までです。"
                        f"解析結果は{len(merged_df)}件でした。証憑または取引数を減らすか、上位プランをご利用ください。"
                    )
                    return
                st.session_state["result"] = all_results
                st.session_state["df"] = merged_df
            except Exception as exc:
                st.error(f"エラーが発生しました：{exc}")
                return

    with right:
        if "df" not in st.session_state:
            st.markdown(
                """
                <div class="app-panel-muted">
                    <div class="eyebrow">RESULT</div>
                    <div class="panel-title">仕訳候補はここに表示されます</div>
                    <p class="panel-copy">証憑画像または取引内容を入力して、基本記帳を始めてください。</p>
                </div>
                """,
                unsafe_allow_html=True,
            )
            return
        df = st.session_state["df"]
        total_amount = int(pd.to_numeric(df["税込金額"], errors="coerce").fillna(0).sum()) if not df.empty else 0
        c1, c2, c3 = st.columns(3)
        c1.metric("取引件数", f"{len(df)} 件")
        c2.metric("税込合計", f"{total_amount:,} 円")
        c3.metric("確認待ち", f"{len(df[df['ステータス'].astype(str).str.contains('確認', na=False)])} 件")

        edited_df = st.data_editor(df, num_rows="fixed", width="stretch", hide_index=True)
        st.session_state["df"] = edited_df

        saved_ledger = load_customer_ledger(customer)
        s1, s2 = st.columns(2)
        s1.metric("保存済み帳簿", f"{len(saved_ledger)} 件")
        if s2.button("この仕訳を保存済み帳簿に追加", width="stretch"):
            total_saved = append_customer_ledger(customer, st.session_state["df"])
            st.success(f"保存しました。保存済み帳簿は合計 {total_saved} 件です。月次決算・年次決算で利用できます。")

        with st.expander("保存済み帳簿の管理"):
            saved_ledger = load_customer_ledger(customer)
            st.dataframe(saved_ledger, width="stretch", hide_index=True)
            opening_balances = load_customer_opening_balances(customer)
            ledger_excel = build_excel(saved_ledger, opening_balances) if not saved_ledger.empty else b""
            if not saved_ledger.empty:
                st.download_button(
                    "保存済み帳簿をExcelでダウンロード",
                    data=ledger_excel,
                    file_name=f"saved_ledger_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                )
            if st.button("保存済み帳簿をクリア", width="stretch"):
                clear_customer_ledger(customer)
                st.success("保存済み帳簿をクリアしました。")
                st.rerun()

        excel_bytes = build_excel(st.session_state["df"], load_customer_opening_balances(customer))
        filename = f"bookkeeping_entries_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        st.download_button(
            "記帳・決算Excelをダウンロード",
            data=excel_bytes,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )


def render_opening_balance_workspace(customer: dict) -> None:
    st.markdown(
        """
        <div class="app-panel">
            <div class="eyebrow">OPENING BALANCE</div>
            <div class="panel-title">使用開始前の財務データを登録</div>
            <p class="panel-copy">このアプリを使い始める前の貸借対照表残高、または当期途中までのPL累計額を入力します。</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    opening_df = load_customer_opening_balances(customer)

    upload_col, guide_col = st.columns([0.9, 1.1], gap="large")
    with upload_col:
        uploaded_file = st.file_uploader(
            "開始残高 CSV / Excel",
            type=["csv", "xlsx", "xls"],
            key="opening_balance_upload",
        )
        if uploaded_file is not None:
            try:
                uploaded_df = read_raw_table_upload(uploaded_file)
                if set(OPENING_BALANCE_COLUMNS).intersection(uploaded_df.columns):
                    imported_df = normalize_opening_balance_df(uploaded_df)
                else:
                    account_column = next((column for column in ["勘定科目", "科目", "項目"] if column in uploaded_df.columns), None)
                    amount_column = next((column for column in ["開始残高", "残高", "金額"] if column in uploaded_df.columns), None)
                    if not account_column:
                        raise ValueError("科目列が見つかりません。'科目' または '勘定科目' 列を含めてください。")
                    imported_df = pd.DataFrame({
                        "科目": uploaded_df[account_column] if account_column else "",
                        "開始残高": uploaded_df[amount_column] if amount_column else 0,
                    })
                    imported_df = normalize_opening_balance_df(imported_df)
                save_customer_opening_balances(customer, imported_df)
                st.success("開始残高を取り込みました。")
                st.rerun()
            except Exception as exc:
                st.error(f"開始残高ファイルを読み取れませんでした：{exc}")

    with guide_col:
        st.info(
            "入力ルール：資産・費用はプラス、負債・純資産・収益も通常の金額をプラスで入力してください。"
            "例：現金 100,000、借入金 500,000、売上高 1,200,000。"
        )

    edited_df = st.data_editor(
        opening_df,
        num_rows="dynamic",
        width="stretch",
        hide_index=True,
        column_config={
            "開始残高": st.column_config.NumberColumn("開始残高", step=1000, format="%d"),
            "分類": st.column_config.SelectboxColumn("分類", options=["資産", "負債", "純資産", "収益", "費用", "未分類"]),
            "帳票": st.column_config.SelectboxColumn("帳票", options=["BS", "PL", "要確認"]),
        },
    )

    b1, b2, b3 = st.columns(3)
    if b1.button("開始残高を保存", type="primary", width="stretch"):
        save_customer_opening_balances(customer, edited_df)
        st.success("保存しました。基本記帳・月次決算・年次決算に反映されます。")
        st.rerun()
    if b2.button("標準科目でリセット", width="stretch"):
        save_customer_opening_balances(customer, build_opening_balance_template())
        st.success("標準科目にリセットしました。")
        st.rerun()

    non_zero = normalize_opening_balance_df(edited_df)
    non_zero = non_zero[non_zero["開始残高"] != 0]
    b3.metric("登録済み残高科目", f"{len(non_zero)} 件")


def render_closing_workspace(period_label: str, transaction_limit: int, customer: dict) -> None:
    st.markdown(
        f"""
        <div class="app-panel">
            <div class="eyebrow">{period_label.upper()}</div>
            <div class="panel-title">顧客の帳簿データから決算表・財務諸表を作成</div>
            <p class="panel-copy">CSV / Excel の仕訳帳をアップロードするか、基本記帳で保存した帳簿を利用できます。</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    source = st.radio("データ元", ["帳簿ファイルをアップロード", "保存済み帳簿を使う"], horizontal=True, key=f"{period_label}_source")
    ledger_df = pd.DataFrame(columns=EXCEL_COLUMNS)

    if source == "帳簿ファイルをアップロード":
        uploaded_files = st.file_uploader(
            "仕訳帳 CSV / Excel（複数ファイル可）",
            type=["csv", "xlsx", "xls"],
            accept_multiple_files=True,
            key=f"{period_label}_ledger_upload",
        )
        if not uploaded_files:
            st.info("顧客の仕訳帳ファイルをアップロードしてください。年次決算では月別ファイルを複数選択できます。")
            return
        try:
            ledger_dfs = []
            for uploaded_file in uploaded_files:
                uploaded_df = read_accounting_upload(uploaded_file)
                if not uploaded_df.empty:
                    uploaded_df["元ファイル名"] = uploaded_file.name
                    ledger_dfs.append(uploaded_df)
            ledger_df = pd.concat(ledger_dfs, ignore_index=True) if ledger_dfs else pd.DataFrame(columns=EXCEL_COLUMNS)
            st.success(f"{len(uploaded_files)}ファイルを読み込みました。合計 {len(ledger_df)} 件の仕訳を集計します。")
        except Exception as exc:
            st.error(f"ファイルを読み取れませんでした：{exc}")
            return
    else:
        ledger_df = load_customer_ledger(customer)
        if ledger_df.empty:
            st.info("保存済み帳簿がありません。基本記帳で仕訳を保存するか、帳簿ファイルをアップロードしてください。")
            return
        st.success(f"保存済み帳簿 {len(ledger_df)} 件を読み込みました。")

    if ledger_df.empty:
        st.warning("決算表を作成できるデータがありません。")
        return
    if len(ledger_df) > transaction_limit and period_label == "月次決算":
        st.warning(f"現在のプランの1回処理目安は{transaction_limit}取引です。大きな帳簿は上位プランをご検討ください。")

    prepared_df = prepare_closing_dataframe(ledger_df)
    selectable_periods = sorted(prepared_df["月"].dropna().unique()) if period_label == "月次決算" else sorted(prepared_df["年"].dropna().unique())
    if selectable_periods:
        selected_period = st.selectbox("対象期間", selectable_periods, index=len(selectable_periods) - 1)
        if period_label == "月次決算":
            ledger_df = prepared_df[prepared_df["月"] == selected_period][EXCEL_COLUMNS].copy()
        else:
            ledger_df = prepared_df[prepared_df["年"] == selected_period][EXCEL_COLUMNS].copy()

    opening_balances = load_customer_opening_balances(customer)
    trial_balance = build_trial_balance(ledger_df, opening_balances)
    profit_and_loss = build_profit_and_loss(trial_balance)
    balance_sheet = build_balance_sheet(trial_balance)

    total_debit = int(trial_balance["借方合計"].sum()) if not trial_balance.empty else 0
    total_credit = int(trial_balance["貸方合計"].sum()) if not trial_balance.empty else 0
    net_income_row = profit_and_loss[profit_and_loss["科目"] == "当期利益"]
    net_income = int(net_income_row["金額"].iloc[0]) if not net_income_row.empty else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("取引件数", f"{len(ledger_df)} 件")
    m2.metric("借方合計", f"{total_debit:,} 円")
    m3.metric("貸方合計", f"{total_credit:,} 円")
    m4.metric("当期利益", f"{net_income:,} 円")

    trial_tab, pl_tab, bs_tab, check_tab = st.tabs(["試算表", "損益計算書", "貸借対照表", "チェックリスト"])
    with trial_tab:
        st.dataframe(trial_balance, width="stretch", hide_index=True)
    with pl_tab:
        st.dataframe(profit_and_loss, width="stretch", hide_index=True)
    with bs_tab:
        st.dataframe(balance_sheet, width="stretch", hide_index=True)
    with check_tab:
        checklist = build_checklist(MONTHLY_CHECKLIST_ITEMS if period_label == "月次決算" else YEARLY_CHECKLIST_ITEMS, period_label)
        st.dataframe(checklist, width="stretch", hide_index=True)

    excel_bytes = build_financial_statement_excel(
        ledger_df,
        trial_balance,
        profit_and_loss,
        balance_sheet,
        period_label,
        load_customer_opening_balances(customer),
    )
    filename = f"{period_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        "決算表・財務諸表Excelをダウンロード",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        width="stretch",
    )


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, layout="wide", initial_sidebar_state="expanded")
    st.markdown(PAGE_CSS, unsafe_allow_html=True)
    st.markdown(
        """
        <div class="app-header">
            <div class="app-title">AI仕訳アシスタント</div>
            <p class="app-subtitle">日本の中小企業・個人事業主向け。証憑画像または取引メモから仕訳候補を作成します。</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    customer = get_logged_in_customer()
    if not customer:
        render_login()

    with st.sidebar:
        st.header("アカウント")
        st.markdown(f"**{customer['name']}**")
        st.caption(customer["username"])
        if st.button("ログアウト", width="stretch"):
            logout_customer()

        st.divider()
        model = DEFAULT_MODEL
        transaction_limit = int(customer["transaction_limit"])
        st.markdown(f"<span class=\"plan-pill\">{customer['plan_label']}</span>", unsafe_allow_html=True)
        st.caption(f"上限：{transaction_limit} 取引/回")
        st.caption(f"AIモデル：{model}")
        st.divider()
        st.info(get_config_value("UPGRADE_CONTACT", "上位プランをご希望の場合は管理者までお問い合わせください。"))

    bookkeeping_tab, opening_tab, monthly_tab, yearly_tab = st.tabs(["基本記帳", "開始残高設定", "月次決算", "年次決算"])
    with bookkeeping_tab:
        render_bookkeeping_workspace(customer, model, transaction_limit)
    with opening_tab:
        render_opening_balance_workspace(customer)
    with monthly_tab:
        render_closing_workspace("月次決算", transaction_limit, customer)
    with yearly_tab:
        render_closing_workspace("年次決算", transaction_limit, customer)


if __name__ == "__main__":
    main()
